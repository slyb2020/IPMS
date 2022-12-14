import pandas as pd
import openpyxl
import wx
import wx.grid as gridlib
import numpy as np

def GetOrderIDFromExcelFile(fileName):
    sheetName = GetSheetNameListFromExcelFileName(fileName)[0]
    data = GetSheetDataFromExcelFileName(fileName,sheetName)
    for i, row in enumerate(data):
        if "Project" in row:
            row = list(row)
            col = row.index("Project")
            for j, item in enumerate(row[col + 1:]):
                if item:
                    return int(item)
    return -1

def GetSubOrderIDListFromExcelFile(fileName):
    sheetNameList = GetSheetNameListFromExcelFileName(fileName)
    result = []
    for sheetName in sheetNameList:
        data = GetSheetDataFromExcelFileName(fileName,sheetName)
        for i, row in enumerate(data):
            if "SubProject" in row:
                row = list(row)
                col = row.index("SubProject")
                for j, item in enumerate(row[col + 1:]):
                    if item:
                        result.append(str(int(item)))
                        break
                break
    return result

def GetSheetNameListFromExcelFileName(fileName):
    wb = openpyxl.load_workbook(fileName)
    sheets = wb.worksheets
    result=[]
    for sheet in sheets:
        result.append(sheet.title)
    return result

def GetSheetDataFromExcelFileName(fileName,sheetName):
    wb = openpyxl.load_workbook(fileName)
    ws = wb.get_sheet_by_name(sheetName)
    data = []
    for row in ws.values:
        temp = []
        for value in row:
            temp.append(value)
        data.append(temp)
    data = np.array(data)
    return data

class Excel2DB():
    def __init__(self, excelFileName, tableName):
        self.excelFileName = excelFileName
        self.tableName = tableName
        self.sheetNameList = self.GetSheetNameListFromExcelFileName()
        self.data = self.GetSheetDataFromExcelFileName(self.sheetNameList[0])[1:]
        # data=[]
        # for i in self.data:
        #     temp=[]
        #     for j in i:
        #         temp.append(str(j))
        #     data.append(temp)
        # self.data=data
        print("data:",self.data)
        self.CreatePriceTable()
        # self.CreateStaffTable()
        # self.EditStaffRecord()

    def GetSheetNameListFromExcelFileName(self):
        wb = openpyxl.load_workbook(self.excelFileName)
        sheets = wb.worksheets
        result = []
        for sheet in sheets:
            result.append(sheet.title)
        return result

    def GetSheetDataFromExcelFileName(self, sheetName):
        wb = openpyxl.load_workbook(self.excelFileName)
        ws = wb.get_sheet_by_name(sheetName)
        data = []
        for row in ws.values:
            temp = []
            for value in row:
                temp.append(value)
            data.append(temp)
        data = np.array(data)
        return data

    def CreatePriceTable(self):
        import pymysql as MySQLdb
        print(dbHostName[whichDB],dbUserName[whichDB],dbPassword[whichDB],dbName[whichDB])
        try:
            db = MySQLdb.connect(host="%s" % dbHostName[whichDB], user='%s' % dbUserName[whichDB],
                                 passwd='%s' % dbPassword[whichDB], db='%s' % dbName[whichDB], charset='utf8')
        except:
            print("????????????%s!" % dbName[whichDB])
            if log:
                log.WriteText("????????????%s!" % dbName[whichDB], colour=wx.RED)
        cursor = db.cursor()
        for data in self.data:
            sql = "INSERT INTO `%s` (`????????????`,`????????????`,`??????????????????`,`????????????`,`????????????`,`????????????`,`??????`,`????????????`,`5000?????????`,`8000?????????`,`10000?????????`,`20000?????????`,`30000?????????`,`40000?????????`,`????????????`)" \
                  "VALUES ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','2022-10-09')" \
                  % (self.tableName, data[1], data[2], data[3], data[4], data[5], data[6], data[7], data[8], data[9], data[10], data[11], data[12], data[13], data[14])
            # sql = "INSERT INTO `%s` (`????????????`,`????????????`,`??????????????????`,`????????????`,`????????????`,`????????????`,`??????`,`SQM Per PIece`,`X?????????`,`Y?????????`)" \
            #       "VALUES ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" \
            #       % (self.tableName, data[0], data[1], data[2], str(data[3]), str(data[4]), str(data[5]), data[6], str(data[7]),
            #          str(data[8]), str(data[9]))
            try:
                cursor.execute(sql)
                db.commit()  # ????????????????????????????????????????????????
            except:
                db.rollback()
                print("error")
        db.close()

    def CreateStaffTable(self):
        import pymysql as MySQLdb
        print(dbHostName[whichDB],dbUserName[whichDB],dbPassword[whichDB],dbName[whichDB])
        try:
            db = MySQLdb.connect(host="%s" % dbHostName[whichDB], user='%s' % dbUserName[whichDB],
                                 passwd='%s' % dbPassword[whichDB], db='%s' % dbName[whichDB], charset='utf8')
        except:
            print("????????????%s!" % dbName[whichDB])
            if log:
                log.WriteText("????????????%s!" % dbName[whichDB], colour=wx.RED)
        cursor = db.cursor()
        for data in self.data:
            psw = str(data[7])[-6:]
            temp = str(data[7])
            birthday = temp[6:10]+'-'+temp[10:12]+'-'+temp[12:14]
            # sql = "INSERT INTO `%s` (`????????????`,`????????????`,`??????????????????`,`????????????`,`????????????`,`????????????`,`??????`,`????????????`,`5000?????????`,`8000?????????`,`10000?????????`,`20000?????????`,`30000?????????`,`40000?????????`,`????????????`)" \
            #       "VALUES ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','2022-10-09')" \
            #       % (self.tableName, data[1], data[2], data[3], data[4], data[5], data[6], data[7], data[8], data[9], data[10], data[11], data[12], data[13], data[14])
            sql = "INSERT INTO `%s` (`???`,`???`,`?????????`,`?????????`,`??????`,`????????????`,`???????????????`,`??????`,`????????????`,`????????????`,`??????`,`??????`)" \
                  "VALUES ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','','%s')" \
                  % (self.tableName, str(data[1]), str(data[2]), str(data[3]), str(data[4]), str(data[5]), str(data[6]), str(data[7]), str(data[8]), str(data[9]), birthday, psw)
            # sql = "INSERT INTO `%s` (`????????????`,`????????????`,`??????????????????`,`????????????`,`????????????`,`????????????`,`??????`,`SQM Per PIece`,`X?????????`,`Y?????????`)" \
            #       "VALUES ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" \
            #       % (self.tableName, data[0], data[1], data[2], str(data[3]), str(data[4]), str(data[5]), data[6], str(data[7]),
            #          str(data[8]), str(data[9]))
            try:
                cursor.execute(sql)
                db.commit()  # ????????????????????????????????????????????????
            except:
                db.rollback()
                print("error")
        db.close()

    def EditStaffRecord(self):
        import pymysql as MySQLdb
        print(dbHostName[whichDB],dbUserName[whichDB],dbPassword[whichDB],dbName[whichDB])
        try:
            db = MySQLdb.connect(host="%s" % dbHostName[whichDB], user='%s' % dbUserName[whichDB],
                                 passwd='%s' % dbPassword[whichDB], db='%s' % dbName[whichDB], charset='utf8')
        except:
            print("????????????%s!" % dbName[whichDB])
            if log:
                log.WriteText("????????????%s!" % dbName[whichDB], colour=wx.RED)
        cursor = db.cursor()
        sql="select `Index`,`??????`,`????????????`,`???????????????`,`????????????` from `info_staff` "
        cursor.execute(sql)
        result = cursor.fetchall()
        temp=[]
        for item in result:
            temp.append(list(item))
        for item in temp:
            if not item[2]:
                timeStr = str(item[4])
                id = "%03d"%item[0]
                sql = "UPDATE `info_staff` SET `????????????`= '%s' where `Index`= %s " % (timeStr[2:4]+timeStr[5:7]+id,item[0])
                try:
                    cursor.execute(sql)
                    db.commit()  # ????????????????????????????????????????????????
                except:
                    print("?????????????????????????????????")
                    db.rollback()
        db.close()


class ExcelGridShowPanel(gridlib.Grid):  ##, mixins.GridAutoEditMixin):
    def __init__(self, parent, fileName,sheetName):
        gridlib.Grid.__init__(self, parent, -1)
        ##mixins.GridAutoEditMixin.__init__(self)
        self.moveTo = None
        self.Bind(wx.EVT_IDLE, self.OnIdle)
        self.data = GetSheetDataFromExcelFileName(fileName,sheetName)
        self.CreateGrid(self.data.shape[0], self.data.shape[1])#, gridlib.Grid.SelectRows)
        self.EnableEditing(False)
        # simple cell formatting
        for row, rowData in enumerate(self.data):
            for col, colData in enumerate(rowData):
                if colData:
                    self.SetCellValue(row,col,str(colData))


        # self.SetCellValue(0, 0, "First cell")
        # self.SetCellValue(1, 1, "Another cell")
        # self.SetCellValue(2, 2, "Yet another cell")
        # self.SetCellValue(3, 3, "This cell is read-only")
        # self.SetCellFont(0, 0, wx.Font(12, wx.FONTFAMILY_ROMAN, wx.FONTSTYLE_ITALIC, wx.FONTWEIGHT_NORMAL))
        # self.SetCellTextColour(1, 1, wx.RED)
        # self.SetCellBackgroundColour(2, 2, wx.CYAN)
        # self.SetReadOnly(3, 3, True)
        #
        # self.SetCellEditor(5, 0, gridlib.GridCellNumberEditor(1,1000))
        # self.SetCellValue(5, 0, "123")
        # self.SetCellEditor(6, 0, gridlib.GridCellFloatEditor())
        # self.SetCellValue(6, 0, "123.34")
        # self.SetCellEditor(7, 0, gridlib.GridCellNumberEditor())
        #
        # self.SetCellValue(6, 3, "You can veto editing this cell")
        #
        # #self.SetRowLabelSize(0)
        # #self.SetColLabelSize(0)
        #
        # # attribute objects let you keep a set of formatting values
        # # in one spot, and reuse them if needed
        # attr = gridlib.GridCellAttr()
        # attr.SetTextColour(wx.BLACK)
        # attr.SetBackgroundColour(wx.RED)
        # attr.SetFont(wx.Font(10, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        #
        # # you can set cell attributes for the whole row (or column)
        # self.SetRowAttr(5, attr)
        #
        # self.SetColLabelValue(0, "Custom")
        # self.SetColLabelValue(1, "column")
        # self.SetColLabelValue(2, "labels")
        #
        # self.SetColLabelAlignment(wx.ALIGN_LEFT, wx.ALIGN_BOTTOM)
        #
        # #self.SetDefaultCellOverflow(False)
        # #r = gridlib.GridCellAutoWrapStringRenderer()
        # #self.SetCellRenderer(9, 1, r)
        #
        # # overflow cells
        # self.SetCellValue( 9, 1, "This default cell will overflow into neighboring cells, but not if you turn overflow off.");
        # self.SetCellSize(11, 1, 3, 3);
        # self.SetCellAlignment(11, 1, wx.ALIGN_CENTRE, wx.ALIGN_CENTRE);
        # self.SetCellValue(11, 1, "This cell is set to span 3 rows and 3 columns");
        #
        #
        # editor = gridlib.GridCellTextEditor()
        # editor.SetParameters('10')
        # self.SetCellEditor(0, 4, editor)
        # self.SetCellValue(0, 4, "Limited text")
        #
        # renderer = gridlib.GridCellAutoWrapStringRenderer()
        # self.SetCellRenderer(15,0, renderer)
        # self.SetCellValue(15,0, "The text in this cell will be rendered with word-wrapping")


        # test all the events
        self.Bind(gridlib.EVT_GRID_CELL_LEFT_CLICK, self.OnCellLeftClick)
        self.Bind(gridlib.EVT_GRID_CELL_RIGHT_CLICK, self.OnCellRightClick)
        self.Bind(gridlib.EVT_GRID_CELL_LEFT_DCLICK, self.OnCellLeftDClick)
        self.Bind(gridlib.EVT_GRID_CELL_RIGHT_DCLICK, self.OnCellRightDClick)

        self.Bind(gridlib.EVT_GRID_LABEL_LEFT_CLICK, self.OnLabelLeftClick)
        self.Bind(gridlib.EVT_GRID_LABEL_RIGHT_CLICK, self.OnLabelRightClick)
        self.Bind(gridlib.EVT_GRID_LABEL_LEFT_DCLICK, self.OnLabelLeftDClick)
        self.Bind(gridlib.EVT_GRID_LABEL_RIGHT_DCLICK, self.OnLabelRightDClick)

        self.Bind(gridlib.EVT_GRID_COL_SORT, self.OnGridColSort)

        self.Bind(gridlib.EVT_GRID_ROW_SIZE, self.OnRowSize)
        self.Bind(gridlib.EVT_GRID_COL_SIZE, self.OnColSize)

        self.Bind(gridlib.EVT_GRID_RANGE_SELECT, self.OnRangeSelect)
        self.Bind(gridlib.EVT_GRID_CELL_CHANGED, self.OnCellChange)
        self.Bind(gridlib.EVT_GRID_SELECT_CELL, self.OnSelectCell)

        self.Bind(gridlib.EVT_GRID_EDITOR_SHOWN, self.OnEditorShown)
        self.Bind(gridlib.EVT_GRID_EDITOR_HIDDEN, self.OnEditorHidden)
        self.Bind(gridlib.EVT_GRID_EDITOR_CREATED, self.OnEditorCreated)


    def OnCellLeftClick(self, evt):
        # self.log.write("OnCellLeftClick: (%d,%d) %s\n" %
        #                (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()

    def OnCellRightClick(self, evt):
        # self.log.write("OnCellRightClick: (%d,%d) %s\n" %
        #                (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()

    def OnCellLeftDClick(self, evt):
        # self.log.write("OnCellLeftDClick: (%d,%d) %s\n" %
        #                (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()

    def OnCellRightDClick(self, evt):
        # self.log.write("OnCellRightDClick: (%d,%d) %s\n" %
        #                (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()

    def OnLabelLeftClick(self, evt):
        # self.log.write("OnLabelLeftClick: (%d,%d) %s\n" %
        #                (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()

    def OnLabelRightClick(self, evt):
        # self.log.write("OnLabelRightClick: (%d,%d) %s\n" %
        #                (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()

    def OnLabelLeftDClick(self, evt):
        # self.log.write("OnLabelLeftDClick: (%d,%d) %s\n" %
        #                (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()

    def OnLabelRightDClick(self, evt):
        # self.log.write("OnLabelRightDClick: (%d,%d) %s\n" %
        #                (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()

    def OnGridColSort(self, evt):
        # self.log.write("OnGridColSort: %s %s" % (evt.GetCol(), self.GetSortingColumn()))
        self.SetSortingColumn(evt.GetCol())

    def OnRowSize(self, evt):
        # self.log.write("OnRowSize: row %d, %s\n" %
        #                (evt.GetRowOrCol(), evt.GetPosition()))
        evt.Skip()

    def OnColSize(self, evt):
        # self.log.write("OnColSize: col %d, %s\n" %
        #                (evt.GetRowOrCol(), evt.GetPosition()))
        evt.Skip()

    def OnRangeSelect(self, evt):
        if evt.Selecting():
            msg = 'Selected'
        else:
            msg = 'Deselected'
        # self.log.write("OnRangeSelect: %s  top-left %s, bottom-right %s\n" %
        #                    (msg, evt.GetTopLeftCoords(), evt.GetBottomRightCoords()))
        evt.Skip()


    def OnCellChange(self, evt):
        # self.log.write("OnCellChange: (%d,%d) %s\n" %
        #                (evt.GetRow(), evt.GetCol(), evt.GetPosition()))

        # Show how to stay in a cell that has bad data.  We can't just
        # call SetGridCursor here since we are nested inside one so it
        # won't have any effect.  Instead, set coordinates to move to in
        # idle time.
        value = self.GetCellValue(evt.GetRow(), evt.GetCol())

        if value == 'no good':
            self.moveTo = evt.GetRow(), evt.GetCol()


    def OnIdle(self, evt):
        if self.moveTo is not None:
            self.SetGridCursor(self.moveTo[0], self.moveTo[1])
            self.moveTo = None

        evt.Skip()


    def OnSelectCell(self, evt):
        if evt.Selecting():
            msg = 'Selected'
        else:
            msg = 'Deselected'
        # self.log.write("OnSelectCell: %s (%d,%d) %s\n" %
        #                (msg, evt.GetRow(), evt.GetCol(), evt.GetPosition()))

        # Another way to stay in a cell that has a bad value...
        row = self.GetGridCursorRow()
        col = self.GetGridCursorCol()

        if self.IsCellEditControlEnabled():
            self.HideCellEditControl()
            self.DisableCellEditControl()

        value = self.GetCellValue(row, col)

        if value == 'no good 2':
            return  # cancels the cell selection

        evt.Skip()


    def OnEditorShown(self, evt):
        if evt.GetRow() == 6 and evt.GetCol() == 3 and \
           wx.MessageBox("Are you sure you wish to edit this cell?",
                        "Checking", wx.YES_NO) == wx.NO:
            evt.Veto()
            return

        # self.log.write("OnEditorShown: (%d,%d) %s\n" %
        #                (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()


    def OnEditorHidden(self, evt):
        if evt.GetRow() == 6 and evt.GetCol() == 3 and \
           wx.MessageBox("Are you sure you wish to  finish editing this cell?",
                        "Checking", wx.YES_NO) == wx.NO:
            evt.Veto()
            return

        # self.log.write("OnEditorHidden: (%d,%d) %s\n" %
        #                (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()


    def OnEditorCreated(self, evt):
        pass
        # self.log.write("OnEditorCreated: (%d, %d) %s\n" %
        #                (evt.GetRow(), evt.GetCol(), evt.GetControl()))

def MakeProductUnitPriceDB():
    temp = GetSheetDataFromExcelFileName('Total.xlsx', '????????? ?????????????????????')
    temp = temp[1:34]
    result = []
    for a in temp:
        a = a[2:12]
        result.append(list(a))
    import pymysql as MySQLdb
    try:
        db = MySQLdb.connect(host="%s" % dbHostName[whichDB], user='%s' % dbUserName[whichDB],
                             passwd='%s' % dbPassword[whichDB], db='%s' % dbName[whichDB], charset='utf8')
    except:
        wx.MessageBox("????????????%s!" % dbName[whichDB], "????????????")
        if log:
            log.WriteText("????????????%s!" % dbName[whichDB], colour=wx.RED)
    cursor = db.cursor()
    for data in result:
        sql = "INSERT INTO ????????????????????? (`????????????`,`????????????`,`??????????????????`,`????????????`,`????????????`,`????????????`,`??????`,`SQM Per PIece`,`X?????????`,`Y?????????`)" \
              "VALUES ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"\
              % (data[0],data[1],data[2],str(data[3]),str(data[4]),str(data[5]),data[6],str(data[7]),str(data[8]),str(data[9]))
        try:
            cursor.execute(sql)
            db.commit()  # ????????????????????????????????????????????????
        except:
            db.rollback()
            print("error")
    db.close()

def MakeProductLaborUnitPriceDB():
    temp = GetSheetDataFromExcelFileName('Total.xlsx', '?????????')
    temp = temp[2:35]
    result = []
    for a in temp:
        a = a[0:3]
        result.append(list(a))
    import pymysql as MySQLdb
    try:
        db = MySQLdb.connect(host="%s" % dbHostName[whichDB], user='%s' % dbUserName[whichDB],
                             passwd='%s' % dbPassword[whichDB], db='%s' % dbName[whichDB], charset='utf8')
    except:
        wx.MessageBox("????????????%s!" % dbName[whichDB], "????????????")
        if log:
            log.WriteText("????????????%s!" % dbName[whichDB], colour=wx.RED)
    cursor = db.cursor()
    for data in result:
        sql = "INSERT INTO ????????????????????? (`????????????`,`??????????????????`,`?????????????????????`)" \
              "VALUES ('%s','%s','%s')"\
              % (data[0],data[1],data[2])
        try:
            cursor.execute(sql)
            db.commit()  # ????????????????????????????????????????????????
        except:
            db.rollback()
            print("error")
    db.close()

from ID_DEFINE import *
if __name__ == "__main__":
    whichDB = 3
    log = None
    # MakeProductLaborUnitPriceDB()
    excel2db = Excel2DB(excelFileName="D:\\BaiduNetdiskWorkspace\\2022?????????\\Luka\\?????? 2022.08.17\\??????????????????.xlsx",tableName='??????????????????')
    # excel2db = Excel2DB(excelFileName="D:\\BaiduNetdiskWorkspace\\2022?????????\\Luka\\?????? 2022.08.17\\??????????????????.xlsx",tableName='info_staff')
